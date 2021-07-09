from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import math
import os
import time


def  print1(fd1,fd2,fd3,source,target,namelist,fen=30): #总表路径 保存路径 模板的路径  总表要复制的列  模板对应的列 名字的列 一张分表有多少人
    wb= load_workbook(fd1)
    
    sheets = wb['Sheet1']
    rows=sheets.max_row
   # print(rows)
    tablerowscount=rows-2         #总户数,起始为第三行时候
    sumfen=math.ceil(tablerowscount/fen)
    
    for i in range(0,sumfen):
        m1=i*fen+3         #m1表示起始为第三行时候
        na1=str(sheets[namelist+str(m1)].value) if  sheets[namelist+str(m1)].value else ''
        fen_row=(i+1)*fen+2 #每页的最后一条在总表的位置 +2表示从第三行开始算
        if fen_row<=rows:
            m2=fen_row
        else:
            m2=rows
        na2=str(sheets[namelist+str(m2)].value) if  sheets[namelist+str(m2)].value else ''
        fd2p=os.path.join(fd2,na1+'-'+na2+'.xlsx')
        if os.path.isfile(fd2p):
            tfd2p=fd2p
        else:
            tfd2p=fd3
        #print(fd2p)
        wb2 =load_workbook(tfd2p)
        sheets2 = wb2['评估汇总']
        #print(sheets2.max_row)
        for x in range(m1,m2+1):
            if x > rows:
                break
            for ii,sourcev in enumerate(source):
                temp=sheets[sourcev+str(x)].value
                sheets2[target[ii]+str(x-i*fen)]=temp
                #print(temp)
        wb2.save(fd2p)
        #pr=[]
        #pr1=[]
       # for row in rows:
           # pr.append(row[2].value)
        #print(pr)


if __name__=='__main__':
    t=time.time()
    lst1=['A','B','C','D','E','F']
    lst2=['A','B','C','D','F','G']
    print1(r'E:\我的文档\vba\金陵东路64号地块给评估公司清册.xlsm',r'E:\我的文档\vba',r'E:\我的文档\vba\杨浦\新模板8.xlsm',lst1,lst2,'B',100)

    print(time.time()-t)

